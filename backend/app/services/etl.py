"""ETL pipeline — raw Excel → tidy Parquet snapshots.

Run with:  make data           (or: python -m app.services.etl)

Inputs (under backend/app/data/raw/, gitignored):
- UK DATA.xlsx                              (sheets: DATABASE, MaterialData, CUSTOMERS)
- Damm Trade Plan - promotions.xlsx         (sheets: Tesco, Sainsbury's, Waitrose, Morrisons, Asda)

Outputs (under backend/app/data/snapshots/, also gitignored):
- wide_monthly.parquet     SKU × SubChannel × month — the primary training table
- targets.parquet          target_hl per SKU × sub_channel × month (derived from
                            prior-year actuals; replaces the misleading null-Hl
                            "budget" rows — those are accounting noise, not a plan)
- promos.parquet           one row per (channel, sku, iso_week), with structured
                            promo_type classified from cell content
- meta.json                brand / SKU / sub_channel / period lists for FE filters

Pipeline steps live in module-level functions so they're individually
testable. The CLI entry point is `main()` at the bottom.
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Final

import polars as pl

from app.paths import RAW_DIR as RAW, SNAPSHOTS_DIR as SNAPSHOTS
from app.services.anonymize import anonymize, anonymize_promo_sheet

# ────────────────────────────────────────────────────────────────────────────
# Paths
# ────────────────────────────────────────────────────────────────────────────

SALES_XLSX = RAW / "UK DATA.xlsx"
PROMO_XLSX = RAW / "Damm Trade Plan - promotions.xlsx"

ALLOWED_SUB_CHANNELS: Final[set[str]] = {
    "GROCERY",
    "FREE TRADE CMBC",
    "NATIONAL ON TRADE",
    "FREE TRADE",
    "CONVENIENCE & WHOLESALE",
    "MDD COPACKING",
}

SPA_MONTH: Final[dict[str, int]] = {
    "Ene": 1, "Feb": 2, "Mar": 3, "Abr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Ago": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dic": 12,
}

# Promo type taxonomy — derived from observation of every cell value in all
# 5 retailer sheets (see audit in commit message). Each type is mutually
# exclusive and grounded in cell content, not heuristic guesses.
PROMO_TYPES: Final[set[str]] = {
    "regular",      # bare number — normal shelf price, not a promo
    "multi-buy",    # "X for £Y", "MTB ...", "2f£Y"
    "price-cut",    # price below this SKU's baseline median by ≥10%
    "rollback",     # "RB £X"
    "clearance",    # "WIGIG £X" — when-it's-gone-it's-gone
    "listing",      # "LAUNCH", "SKU replacement"
    "no-listing",   # cell empty — SKU not stocked that week
}


# ────────────────────────────────────────────────────────────────────────────
# Pure helpers (testable without I/O)
# ────────────────────────────────────────────────────────────────────────────

def parse_period(s: str | None) -> date | None:
    """`Abr.25` → date(2025, 4, 1). Anything unparseable → None."""
    if not s:
        return None
    try:
        m, y = s.split(".")
        return date(2000 + int(y), SPA_MONTH[m[:3]], 1)
    except (KeyError, ValueError):
        return None


def extract_cliente(s: str | None) -> int | None:
    """`1/1/91/117738 CARLSBERG SUPPLY COMPANY AG` → 117738."""
    if not s:
        return None
    try:
        last = str(s).split("/")[-1].strip()
        return int(last.split()[0])
    except (ValueError, IndexError):
        return None


def extract_material(s: str | None) -> str | None:
    """`K015600 CERVEZA CORRIENTE EXPORT DAMM A21` → `K015600`."""
    if not s:
        return None
    parts = str(s).strip().split()
    return parts[0] if parts else None


def extract_customer_name(s: str | None) -> str | None:
    if not s:
        return None
    try:
        last = str(s).split("/")[-1].strip()
        parts = last.split(None, 1)
        return parts[1].strip() if len(parts) == 2 else None
    except (ValueError, IndexError):
        return None


# ────────────────────────────────────────────────────────────────────────────
# Loaders — sales side
# ────────────────────────────────────────────────────────────────────────────

def load_sales_raw() -> pl.DataFrame:
    """Load DATABASE sheet, parse periods, extract numeric IDs."""
    df = pl.read_excel(SALES_XLSX, sheet_name="DATABASE")
    print(f"  · DATABASE: {len(df):,} rows, {len(df.columns)} cols")
    out = df.with_columns(
        pl.col("AÑO CALENDARIO").map_elements(parse_period, return_dtype=pl.Date).alias("date"),
        pl.col("Cod. Cliente").map_elements(extract_cliente, return_dtype=pl.Int64).alias("cliente_id"),
        pl.col("Cod. Material").map_elements(extract_material, return_dtype=pl.String).alias("material_id"),
        pl.col("Cod. Cliente").map_elements(extract_customer_name, return_dtype=pl.String).alias("customer_name_raw"),
    )
    out = out.filter(pl.col("date").is_not_null())
    print(f"  · after period parse: {len(out):,} rows ({out['date'].min()} → {out['date'].max()})")
    return out


def load_customers_uk() -> pl.DataFrame:
    df = pl.read_excel(SALES_XLSX, sheet_name="CUSTOMERS")
    uk = df.filter(pl.col("Pais") == "Reino Unido")
    sub_values = set(uk["SubChannel"].drop_nulls().unique().to_list())
    unexpected = sub_values - ALLOWED_SUB_CHANNELS
    if unexpected:
        print(f"  ! unexpected SubChannel values (will be dropped from final): {unexpected}")
    return uk.with_columns(
        pl.col("Agrupacion BU3").map_elements(anonymize, return_dtype=pl.String).alias("customer_anon"),
    ).select(
        pl.col("Cod. Cliente").alias("cliente_id"),
        pl.col("Sales Channel").alias("sales_channel"),
        pl.col("SubChannel").alias("sub_channel"),
        pl.col("Agrupacion BU3").alias("agrupacion_raw"),
        "customer_anon",
    )


def load_materials() -> pl.DataFrame:
    df = pl.read_excel(SALES_XLSX, sheet_name="MaterialData")
    keep = [c for c in df.columns if not c.startswith("Unnamed")]
    df = df.select(keep)
    out = df.with_columns(
        pl.col("Cod. Material").cast(pl.String).str.strip_chars().alias("material_id"),
    ).select(
        "material_id",
        pl.col("Marca").alias("brand"),
        pl.col("Línea Negocio").alias("business_line"),
        pl.col("Tipo Envase").alias("package_type_es"),
        pl.col("PACK TYPE").alias("pack_type"),
        pl.col("PACK SIZE").alias("pack_size"),
        pl.col("ALC. %").alias("alc_pct"),
        pl.col("L por SKU").alias("litres_per_sku"),
    ).unique(subset=["material_id"], keep="first")

    # Construct a human-readable label per SKU.
    # Format: "Estrella Damm · 330ml Can" — much friendlier than "EX23SRAN".
    # Falls back gracefully when fields are missing.
    def _title_case_brand(b: str | None) -> str:
        if not b: return "(unknown brand)"
        words = b.lower().split()
        return " ".join(w.capitalize() for w in words)

    out = out.with_columns(
        label=pl.struct(["brand", "pack_size", "pack_type"]).map_elements(
            lambda s: _build_sku_label(s["brand"], s["pack_size"], s["pack_type"]),
            return_dtype=pl.String,
        ),
    )
    return out


def _build_sku_label(brand: str | None, pack_size: str | None, pack_type: str | None) -> str:
    """Compose a friendly SKU label: 'Estrella Damm · 330ml Can'.

    Token-dedupe approach: lowercase pack_size + pack_type, split into
    tokens, keep only the first occurrence of each. Then re-case sensibly.
    Examples:
      '660ML NR' + 'NR BOTTLE'  → '660ml nr bottle'
      '50L KEG'  + 'KEG'        → '50l keg'
      '330ML CAN' + 'CAN'       → '330ml can'
      '1/3 SR.'  + 'CAN'        → '1/3 sr. can'
    """
    brand_label = (
        " ".join(w.capitalize() for w in brand.lower().split())
        if brand else "Unknown brand"
    )
    raw_tokens: list[str] = []
    for src in (pack_size, pack_type):
        if not src or src == "Dummy":
            continue
        raw_tokens.extend(src.strip().lower().split())
    # Preserve order, dedupe consecutive repeats and global repeats
    seen: set[str] = set()
    pack_tokens: list[str] = []
    for t in raw_tokens:
        if t and t not in seen:
            pack_tokens.append(t)
            seen.add(t)
    parts = [brand_label]
    if pack_tokens:
        parts.append(" ".join(pack_tokens))
    return " · ".join(parts)


# ────────────────────────────────────────────────────────────────────────────
# Sales transforms
# ────────────────────────────────────────────────────────────────────────────

def filter_actuals(sales: pl.DataFrame) -> pl.DataFrame:
    """Keep only rows that represent real sales volume.

    Null-Hl rows in this dataset turn out to be accounting adjustments
    (returns, credit notes, fee allocations) spread across all years, NOT a
    planned budget — see DATA.md audit. We drop them.
    """
    actuals = sales.filter(pl.col("Hl").is_not_null())
    n_neg = (actuals["Hl"] < 0).sum()
    print(f"  · null-Hl rows dropped (accounting noise, not budget): {len(sales) - len(actuals):,}")
    print(f"  · negative-Hl rows that will be netted: {n_neg:,}")
    return actuals


def net_returns(actuals: pl.DataFrame) -> pl.DataFrame:
    """Net negative Hl against same (cliente, material, month). Drop where net ≤ 0."""
    netted = (
        actuals
        .group_by(["cliente_id", "material_id", "date"])
        .agg(
            pl.col("Hl").sum(),
            pl.col("Venta Neta").sum().alias("revenue_gbp"),
            pl.col("Margen Bruto").sum().alias("margin_gbp"),
        )
        .filter(pl.col("Hl") > 0)
    )
    print(f"  · after net+filter: {len(netted):,} rows")
    return netted


def join_uk(actuals_netted: pl.DataFrame, customers_uk: pl.DataFrame, materials: pl.DataFrame) -> pl.DataFrame:
    joined = (
        actuals_netted
        .join(customers_uk, on="cliente_id", how="inner")
        .join(materials, on="material_id", how="inner")
        .filter(pl.col("sub_channel").is_in(list(ALLOWED_SUB_CHANNELS)))
    )
    print(f"  · joined UK: {len(joined):,} rows  ·  "
          f"customers={joined['cliente_id'].n_unique()}  "
          f"SKUs={joined['material_id'].n_unique()}  "
          f"brands={joined['brand'].n_unique()}")
    return joined


def aggregate_monthly(joined: pl.DataFrame) -> pl.DataFrame:
    return (
        joined
        .group_by(["material_id", "brand", "sub_channel", "sales_channel", "date"])
        .agg(
            pl.col("Hl").sum(),
            pl.col("revenue_gbp").sum(),
            pl.col("margin_gbp").sum(),
            pl.col("cliente_id").n_unique().alias("n_customers"),
            (pl.col("customer_anon") == "Distributor (B2B)").any().alias("has_cmbc"),
        )
        .with_columns(
            pl.col("date").dt.month().alias("month"),
            pl.col("date").dt.quarter().alias("quarter"),
            pl.col("date").dt.year().alias("year"),
        )
        .sort(["material_id", "sub_channel", "date"])
    )


def derive_future_targets(monthly: pl.DataFrame, horizon_months: int = 9) -> pl.DataFrame:
    """Extend the targets table into the next `horizon_months` for every series.

    For each (material_id, sub_channel) and each future month, target_hl =
    same-month-prior-year actual, or trailing 3-month median if not available.
    """
    last_date = monthly["date"].max()
    future_dates = []
    from datetime import date as date_t
    y, m = last_date.year, last_date.month
    for _ in range(horizon_months):
        m += 1
        if m > 12: m = 1; y += 1
        future_dates.append(date_t(y, m, 1))

    series_keys = monthly.select(["material_id", "sub_channel"]).unique()
    # Cross join series × future_dates
    future_rows = []
    for d in future_dates:
        for k in series_keys.iter_rows(named=True):
            future_rows.append({
                "material_id": k["material_id"],
                "sub_channel": k["sub_channel"],
                "date": d,
            })
    future = pl.DataFrame(future_rows)

    # Same-month-prior-year lookup
    prior_year = monthly.select(
        pl.col("material_id"),
        pl.col("sub_channel"),
        pl.col("date").dt.offset_by("12mo").alias("date"),
        pl.col("Hl").alias("prior_year_hl"),
    )
    future = future.join(prior_year, on=["material_id", "sub_channel", "date"], how="left")

    # Trailing 3-month median per series for cold-start
    trailing_window = (
        monthly.sort(["material_id", "sub_channel", "date"])
        .group_by(["material_id", "sub_channel"], maintain_order=True)
        .agg(
            last_3_median=pl.col("Hl").tail(3).median(),
        )
    )
    future = future.join(trailing_window, on=["material_id", "sub_channel"], how="left")
    future = future.with_columns(
        target_hl=pl.coalesce(["prior_year_hl", "last_3_median"]).fill_null(0.0),
        target_source=pl.when(pl.col("prior_year_hl").is_not_null()).then(pl.lit("prior_year"))
                        .otherwise(pl.lit("trailing_median")),
    ).select(["material_id", "sub_channel", "date", "target_hl", "target_source"])
    return future


def derive_targets(monthly: pl.DataFrame) -> pl.DataFrame:
    """Build a target_hl series per (material, sub_channel, date).

    There is no explicit budget in UK DATA.xlsx — the null-Hl rows turned out
    to be accounting adjustments, not a plan (see DATA.md). For forecast-vs-
    target comparison we derive a target from prior-year same-month actuals:

        target_hl[m] = actual_hl[m - 12]                if m - 12 exists
                     = trailing-3-month median × 1.0    otherwise (cold-start)

    This is explicitly a *derived* target. The dashboard surfaces this clearly
    via target_source ∈ {"prior_year", "trailing_median"}.

    Why prior-year-actuals: in CPG, monthly budgets are typically set as
    "match last year ± growth %". Using prior-year actuals as the target is
    the most defensible baseline absent a real plan. Growth % can be applied
    per-brand later as a config knob.
    """
    base = monthly.select(["material_id", "sub_channel", "date", "Hl"])

    # Prior-year join — date minus 12 months
    prior = base.with_columns(
        target_date=pl.col("date").dt.offset_by("12mo"),
    ).select(
        pl.col("material_id"),
        pl.col("sub_channel"),
        pl.col("target_date").alias("date"),
        pl.col("Hl").alias("prior_year_hl"),
    )

    targets = base.join(prior, on=["material_id", "sub_channel", "date"], how="left")

    # Trailing-3-month median per (material, sub_channel) as cold-start fallback
    trailing = (
        base.sort(["material_id", "sub_channel", "date"])
        .with_columns(
            trailing_med=pl.col("Hl")
                .rolling_median(window_size=3, min_samples=1)
                .over(["material_id", "sub_channel"]),
        )
        .select("material_id", "sub_channel", "date", "trailing_med")
    )

    out = targets.join(trailing, on=["material_id", "sub_channel", "date"], how="left").with_columns(
        target_hl=pl.coalesce(["prior_year_hl", "trailing_med"]),
        target_source=pl.when(pl.col("prior_year_hl").is_not_null())
            .then(pl.lit("prior_year"))
            .otherwise(pl.lit("trailing_median")),
    ).select("material_id", "sub_channel", "date", "target_hl", "target_source")

    n_prior = (out["target_source"] == "prior_year").sum()
    n_trail = (out["target_source"] == "trailing_median").sum()
    print(f"  · targets: {len(out):,} rows · prior_year={n_prior:,} · trailing_median={n_trail:,}")
    return out


# ────────────────────────────────────────────────────────────────────────────
# Promo parsing — per-retailer structural parsers + content-based classifier
# ────────────────────────────────────────────────────────────────────────────

@dataclass
class PromoCell:
    channel: str
    sku: str            # SKU label as written in the retailer's sheet
    iso_week: date      # Monday of the ISO week
    week_number: int | None
    price_gbp: float | None
    multi_buy_offer: str | None      # raw string e.g. "2 for £23"
    cell_kind: str                   # 'price' | 'multi-buy' | 'rollback' | 'clearance' | 'listing' | 'other'
    raw_value: str                   # exact original cell content


# Cell-content classification regex set — exact strings observed in the sheets
RE_MONEY_FOR     = re.compile(r"(\d+)\s*(?:for|f)\s*£\s*([\d.]+)", re.I)
RE_MTB           = re.compile(r"\bMTB\b", re.I)            # "Multi-Tier Buy"
RE_ROLLBACK      = re.compile(r"\b(?:RB|rollback|roll back)\b\s*£?\s*([\d.]+)?", re.I)
RE_CLEARANCE     = re.compile(r"\bWIGIG\b", re.I)          # When-It's-Gone-It's-Gone
RE_LAUNCH        = re.compile(r"\b(LAUNCH|SKU replacement|new listing)\b", re.I)
RE_PRICE         = re.compile(r"^£?\s*\d+(?:\.\d+)?$")


def _classify_cell(value) -> tuple[str, float | None, str | None]:
    """Return (cell_kind, price_gbp, multi_buy_offer) for one cell.

    - bare number → ('price', float, None)
    - "X for £Y"  → ('multi-buy', None, raw string)
    - "MTB ..."   → ('multi-buy', None, raw string)
    - "RB £X"     → ('rollback', float|None, None)
    - "WIGIG £X"  → ('clearance', float|None, None)
    - "LAUNCH"    → ('listing', None, None)
    - other       → ('other', None, None)
    """
    if isinstance(value, (int, float)):
        return ("price", float(value), None)
    s = str(value).strip()
    if not s:
        return ("other", None, None)
    if RE_PRICE.match(s):
        return ("price", float(s.lstrip("£")), None)
    if RE_LAUNCH.search(s):
        return ("listing", None, None)
    if RE_CLEARANCE.search(s):
        m = re.search(r"£\s*([\d.]+)", s)
        return ("clearance", float(m.group(1)) if m else None, s)
    if RE_ROLLBACK.search(s):
        m = re.search(r"£\s*([\d.]+)", s)
        return ("rollback", float(m.group(1)) if m else None, s)
    if RE_MONEY_FOR.search(s) or RE_MTB.search(s):
        return ("multi-buy", None, s)
    return ("other", None, s)


def _iso_monday(d: datetime | date) -> date:
    """Snap any date to the Monday of its ISO week."""
    if isinstance(d, datetime):
        d = d.date()
    return d - timedelta(days=d.weekday())


def _parse_dd_mm_range(s: str, default_year: int) -> tuple[date, date] | None:
    """Parse '01/01-20/01' → (Mon of 2026-01-01, Mon of 2026-01-20)."""
    m = re.match(r"\s*(\d{1,2})/(\d{1,2})\s*[-–]\s*(\d{1,2})/(\d{1,2})\s*", s)
    if not m:
        return None
    d1, mo1, d2, mo2 = (int(x) for x in m.groups())
    try:
        return _iso_monday(date(default_year, mo1, d1)), _iso_monday(date(default_year, mo2, d2))
    except ValueError:
        return None


def _iter_weeks(start: date, end: date):
    """Yield ISO-Monday dates from start to end inclusive."""
    cur = start
    while cur <= end:
        yield cur
        cur = cur + timedelta(days=7)


# Per-retailer parsers — each understands its own sheet's exact structure
# (rows/columns identified by manual audit). They all return list[PromoCell].

def _parse_grid_with_week_row(ws, channel: str,
                              week_row_idx: int,
                              date_row_idx: int,
                              sku_label_col: int,
                              first_data_col: int,
                              first_data_row: int) -> list[PromoCell]:
    """
    Generic parser for the common 'wide grid' layout used by Tesco, Sainsbury's,
    and Waitrose:

    - one row of period codes (P1/P13/C1...)
    - one row of week-start datetimes
    - one row of week numbers (optional)
    - SKU label column on the left
    - matrix of (SKU × week) price/promo cells

    Caller passes the exact row indexes (zero-based) after a per-sheet audit.
    """
    rows = list(ws.values)
    if date_row_idx >= len(rows) or week_row_idx >= len(rows):
        return []
    date_row = rows[date_row_idx]
    week_row = rows[week_row_idx] if 0 <= week_row_idx < len(rows) else [None] * len(date_row)

    # Build week-index → (iso_monday, week_number)
    week_columns: list[tuple[int, date, int | None]] = []
    for j, cell in enumerate(date_row):
        if j < first_data_col:
            continue
        if isinstance(cell, (datetime, date)):
            wn = week_row[j] if j < len(week_row) else None
            if isinstance(wn, (int, float)):
                wn = int(wn)
            elif isinstance(wn, str) and wn.strip().isdigit():
                wn = int(wn.strip())
            else:
                wn = None
            week_columns.append((j, _iso_monday(cell), wn))
    if not week_columns:
        return []

    out: list[PromoCell] = []
    for row in rows[first_data_row:]:
        if not row:
            continue
        sku_cell = row[sku_label_col] if sku_label_col < len(row) else None
        if not isinstance(sku_cell, str) or not sku_cell.strip():
            continue
        sku = sku_cell.strip()
        for j, iso_mon, wn in week_columns:
            if j >= len(row):
                continue
            val = row[j]
            if val is None:
                # Empty cell → SKU not listed that week
                out.append(PromoCell(
                    channel=channel, sku=sku, iso_week=iso_mon, week_number=wn,
                    price_gbp=None, multi_buy_offer=None, cell_kind="other", raw_value=""
                ))
                continue
            kind, price, offer = _classify_cell(val)
            out.append(PromoCell(
                channel=channel, sku=sku, iso_week=iso_mon, week_number=wn,
                price_gbp=price, multi_buy_offer=offer, cell_kind=kind,
                raw_value=str(val).strip(),
            ))
    return out


def _parse_tesco(ws) -> list[PromoCell]:
    """Tesco: header rows 0-1 = event labels (ignored — those are retailer
    promo-calendar event tags, not Damm classifications). Period codes row 2,
    week dates row 3, week numbers row 4, SKU label col 2, prices col 3+."""
    return _parse_grid_with_week_row(
        ws, channel="Grocer A",
        week_row_idx=4, date_row_idx=3,
        sku_label_col=2, first_data_col=3, first_data_row=5,
    )


def _parse_sainsburys(ws) -> list[PromoCell]:
    """Sainsbury's: period code row 1, dates row 2, week numbers row 3,
    SKU label col 0, data col 1+."""
    return _parse_grid_with_week_row(
        ws, channel="Grocer B",
        week_row_idx=3, date_row_idx=2,
        sku_label_col=0, first_data_col=1, first_data_row=4,
    )


def _parse_waitrose(ws) -> list[PromoCell]:
    """Waitrose: same structure as Sainsbury's."""
    return _parse_grid_with_week_row(
        ws, channel="Grocer E",
        week_row_idx=3, date_row_idx=2,
        sku_label_col=0, first_data_col=1, first_data_row=4,
    )


def _parse_morrisons(ws) -> list[PromoCell]:
    """Morrisons is pivoted differently — wide, sparse, 'Start/end' row + 'Period' row.
    Row 1 has alternating (start_date, NaT, start_date, NaT, ...) — odd cols are starts
    of period windows, not weekly. Skus start from row 6.

    Because the granularity isn't weekly we treat each period as a single
    promo-window record: one row per (sku, period_start_date)."""
    rows = list(ws.values)
    if len(rows) < 6:
        return []
    date_row = rows[1]
    period_row = rows[2]
    # Collect (col_idx, start_date, period_code) for cells where date_row has a real datetime
    windows: list[tuple[int, date, str | None]] = []
    for j, c in enumerate(date_row):
        if isinstance(c, (datetime, date)):
            pcode = period_row[j] if j < len(period_row) else None
            if isinstance(pcode, str):
                pcode = pcode.strip()
            else:
                pcode = None
            windows.append((j, _iso_monday(c), pcode))
    if not windows:
        return []

    out: list[PromoCell] = []
    for row in rows[6:]:
        if not row:
            continue
        sku = row[0] if row[0] else None
        if not isinstance(sku, str) or not sku.strip():
            continue
        sku = sku.strip()
        for j, iso_mon, pcode in windows:
            if j >= len(row) or row[j] is None:
                continue
            val = row[j]
            kind, price, offer = _classify_cell(val)
            if kind == "other" and not str(val).strip():
                continue
            out.append(PromoCell(
                channel="Grocer D", sku=sku, iso_week=iso_mon, week_number=None,
                price_gbp=price, multi_buy_offer=offer, cell_kind=kind,
                raw_value=str(val).strip(),
            ))
    return out


def _parse_asda(ws) -> list[PromoCell]:
    """Asda uses 'R1'..'R14' codes; row 1 has 'dd/mm-dd/mm' date-range strings.
    The default year is 2026 (inferred from the other sheets which use 2026 dates)."""
    rows = list(ws.values)
    if len(rows) < 3:
        return []
    date_range_row = rows[1]
    windows: list[tuple[int, list[date]]] = []
    for j, c in enumerate(date_range_row):
        if not isinstance(c, str):
            continue
        rng = _parse_dd_mm_range(c, default_year=2026)
        if rng:
            windows.append((j, list(_iter_weeks(rng[0], rng[1]))))
    if not windows:
        return []

    out: list[PromoCell] = []
    for row in rows[2:]:
        if not row:
            continue
        sku = row[0]
        if not isinstance(sku, str) or not sku.strip():
            continue
        sku = sku.strip()
        for j, weeks in windows:
            if j >= len(row) or row[j] is None:
                continue
            val = row[j]
            kind, price, offer = _classify_cell(val)
            for iso_mon in weeks:
                out.append(PromoCell(
                    channel="Grocer C", sku=sku, iso_week=iso_mon, week_number=None,
                    price_gbp=price, multi_buy_offer=offer, cell_kind=kind,
                    raw_value=str(val).strip(),
                ))
    return out


_PARSERS: dict[str, callable] = {
    "Tesco":         _parse_tesco,
    "Sainsbury's":   _parse_sainsburys,
    "Waitrose":      _parse_waitrose,
    "Morrisons":     _parse_morrisons,
    "Asda":          _parse_asda,
}


def parse_promos_all() -> pl.DataFrame:
    """Parse every retailer with its own sheet-specific parser, then promote
    `cell_kind` → `promo_type` by adding the *price-cut* category, which can
    only be determined after we know each SKU's baseline price.
    """
    import openpyxl
    wb = openpyxl.load_workbook(PROMO_XLSX, read_only=True, data_only=True)

    cells: list[PromoCell] = []
    for raw_sheet_name in wb.sheetnames:
        key = raw_sheet_name.strip()
        parser = _PARSERS.get(key)
        if parser is None:
            print(f"  ! no parser for sheet {raw_sheet_name!r} — skipped")
            continue
        ws = wb[raw_sheet_name]
        these = parser(ws)
        print(f"  · {raw_sheet_name:<14} {len(these):>5} cells parsed  →  channel={(these[0].channel if these else '—')}")
        cells.extend(these)

    if not cells:
        return pl.DataFrame(schema={
            "channel": pl.String, "sku": pl.String, "iso_week": pl.Date,
            "week_number": pl.Int32, "price_gbp": pl.Float64,
            "multi_buy_offer": pl.String, "promo_type": pl.String,
            "on_promo": pl.Boolean, "baseline_price_gbp": pl.Float64,
            "raw_value": pl.String,
        })

    df = pl.DataFrame(
        [c.__dict__ for c in cells],
        schema={
            "channel": pl.String, "sku": pl.String, "iso_week": pl.Date,
            "week_number": pl.Int32, "price_gbp": pl.Float64,
            "multi_buy_offer": pl.String, "cell_kind": pl.String,
            "raw_value": pl.String,
        },
    )

    # Baseline price per (channel, sku) = median of all 'price' cells
    baselines = (
        df.filter(pl.col("cell_kind") == "price")
        .group_by(["channel", "sku"])
        .agg(baseline_price_gbp=pl.col("price_gbp").median())
    )
    df = df.join(baselines, on=["channel", "sku"], how="left")

    # Promote cell_kind into a clean promo_type
    df = df.with_columns(
        promo_type=pl.when(pl.col("cell_kind") == "listing").then(pl.lit("listing"))
        .when(pl.col("cell_kind") == "clearance").then(pl.lit("clearance"))
        .when(pl.col("cell_kind") == "rollback").then(pl.lit("rollback"))
        .when(pl.col("cell_kind") == "multi-buy").then(pl.lit("multi-buy"))
        .when((pl.col("cell_kind") == "price")
              & (pl.col("baseline_price_gbp").is_not_null())
              & (pl.col("price_gbp") < pl.col("baseline_price_gbp") * 0.9))
            .then(pl.lit("price-cut"))
        .when(pl.col("cell_kind") == "price").then(pl.lit("regular"))
        .when(pl.col("cell_kind") == "other").then(pl.lit("no-listing"))
        .otherwise(pl.lit("no-listing"))
    ).with_columns(
        on_promo=pl.col("promo_type").is_in(["multi-buy", "price-cut", "rollback", "clearance"]),
    ).drop("cell_kind")

    # Sanity check — every promo_type must be in our taxonomy
    seen_types = set(df["promo_type"].unique().to_list())
    unexpected = seen_types - PROMO_TYPES
    assert not unexpected, f"BUG: unexpected promo_type values produced: {unexpected}"
    return df


# ────────────────────────────────────────────────────────────────────────────
# External enrichment (Phase 1 = UK holidays only; weather/trends in Phase 2)
# ────────────────────────────────────────────────────────────────────────────

def attach_uk_holidays(monthly: pl.DataFrame) -> pl.DataFrame:
    import holidays as hd
    years = sorted(monthly["year"].unique().to_list())
    uk = hd.country_holidays("GB", years=years)
    counts: dict[date, int] = {}
    for d in uk.keys():
        first = date(d.year, d.month, 1)
        counts[first] = counts.get(first, 0) + 1
    return monthly.with_columns(
        pl.col("date").map_elements(lambda d: counts.get(d, 0), return_dtype=pl.Int32).alias("uk_holidays_count")
    )


def attach_external(monthly: pl.DataFrame) -> pl.DataFrame:
    """Join weather, Google Trends, and ONS retail series by month-start date.

    Sources are cached under app/data/cache/. Robust to individual source
    failures — see external.py.
    """
    from app.services.external import fetch_all_external
    date_min = monthly["date"].min()
    date_max = monthly["date"].max()
    ext = fetch_all_external(date_min, date_max)
    out = monthly.join(ext, on="date", how="left")
    # Forward-fill then zero-fill for any external column that's null on a given month
    fill_cols = [c for c in ext.columns if c != "date"]
    out = out.with_columns(*[pl.col(c).fill_null(0.0) for c in fill_cols])
    n_filled = sum((monthly[c].null_count() if c in monthly.columns else 0) for c in fill_cols)
    print(f"  · attached {len(fill_cols)} external columns: {fill_cols}")
    return out


def attach_planned_promos(monthly: pl.DataFrame, promos: pl.DataFrame) -> pl.DataFrame:
    """Per-month planned-promo intensity per (brand × sub_channel).

    Adds two columns to `monthly`:
      n_planned_promos     — # of (sku × week) promo cells with on_promo=true
                              that landed in the month, aggregated to brand.
      avg_planned_discount — mean discount fraction over cells with both
                              price_gbp and baseline_price_gbp set; rows
                              where the discount can't be derived (multi-buy
                              without unit price, clearance without £)
                              still count toward n_planned_promos but
                              don't move the average.

    Brand matching: the trade-plan workbook stores SKUs as retailer labels
    ("Estrella 10x330ml Can") with no link to material_id. We match each
    promo SKU to the brand whose first-word (lowercased) appears in the
    SKU label; ties broken by highest historical GROCERY Hl so multi-brand
    families resolve to the dominant brand (e.g. "Damm Lemon ..." → DAMM
    LEMON rather than the rarer DAMM TOSTADA).

    Sub-channel: promos.parquet only covers grocers (channels Grocer A-E),
    all of which map to sub_channel='GROCERY'. Non-GROCERY rows in
    `monthly` are left at 0 / 0.0.
    """
    agg = aggregate_planned_promos(monthly, promos)
    if agg.height == 0:
        return monthly.with_columns(
            n_planned_promos=pl.lit(0, dtype=pl.Int32),
            avg_planned_discount=pl.lit(0.0, dtype=pl.Float64),
        )

    n_rows_with_promo = int((agg["n_planned_promos"] > 0).sum())
    n_brands = agg["brand"].n_unique()
    print(
        f"  · planned promos: {n_rows_with_promo} brand-month buckets covered "
        f"({n_brands} brands matched)"
    )

    out = monthly.join(
        agg, on=["brand", "sub_channel", "date"], how="left"
    ).with_columns(
        pl.col("n_planned_promos").fill_null(0).cast(pl.Int32),
        pl.col("avg_planned_discount").fill_null(0.0),
    )
    return out


def aggregate_planned_promos(monthly: pl.DataFrame, promos: pl.DataFrame) -> pl.DataFrame:
    """Promo intensity per (brand × sub_channel='GROCERY' × month-start date).

    Shared by ETL (which left-joins onto historical `monthly`) and forecast
    inference (which needs the SAME aggregation projected onto FUTURE months
    the trade plan covers — those months don't exist in wide_monthly). The
    `monthly` arg is only used to compute brand prevalence for tie-breaking.
    """
    if promos.height == 0:
        return pl.DataFrame(schema={
            "brand": pl.String, "sub_channel": pl.String, "date": pl.Date,
            "n_planned_promos": pl.Int32, "avg_planned_discount": pl.Float64,
        })

    # Brand prevalence in GROCERY — drives tie-breaks when multiple brands
    # share the same first word ("DAMM ..." family).
    prevalence_df = (
        monthly.filter(pl.col("sub_channel") == "GROCERY")
        .group_by("brand")
        .agg(pl.col("Hl").sum().alias("hl_total"))
        .sort("hl_total", descending=True)
    )
    prevalence = {r["brand"]: float(r["hl_total"]) for r in prevalence_df.iter_rows(named=True) if r["brand"]}
    real_brands = [b for b in prevalence.keys() if b and b.strip() and b != "#"]

    by_first_word: dict[str, list[tuple[str, str]]] = {}
    for b in real_brands:
        fw = b.lower().split()[0]
        by_first_word.setdefault(fw, []).append((b.lower(), b))
    for fw, lst in by_first_word.items():
        lst.sort(key=lambda x: (-len(x[0]), -prevalence.get(x[1], 0.0)))

    def _match_brand(sku_label: str | None) -> str | None:
        if not sku_label:
            return None
        s = sku_label.lower()
        tokens = s.split()
        if not tokens:
            return None
        candidates = by_first_word.get(tokens[0])
        if not candidates:
            return None
        for full_lower, brand_canonical in candidates:
            if full_lower in s:
                return brand_canonical
        return candidates[0][1]

    df = promos.with_columns(
        pl.col("sku").map_elements(_match_brand, return_dtype=pl.String).alias("_brand"),
        pl.col("iso_week").dt.truncate("1mo").alias("_month"),
    ).filter(pl.col("_brand").is_not_null())

    df = df.with_columns(
        pl.when(
            pl.col("on_promo")
            & pl.col("price_gbp").is_not_null()
            & pl.col("baseline_price_gbp").is_not_null()
            & (pl.col("baseline_price_gbp") > 0)
            & (pl.col("price_gbp") < pl.col("baseline_price_gbp"))
        )
        .then(
            (pl.col("baseline_price_gbp") - pl.col("price_gbp"))
            / pl.col("baseline_price_gbp")
        )
        .otherwise(None)
        .alias("_discount_pct"),
    )

    return (
        df.group_by(["_brand", "_month"])
        .agg(
            n_planned_promos=pl.col("on_promo").cast(pl.Int32).sum(),
            avg_planned_discount=pl.col("_discount_pct").mean(),
        )
        .rename({"_brand": "brand", "_month": "date"})
        .with_columns(
            pl.col("avg_planned_discount").fill_null(0.0),
            sub_channel=pl.lit("GROCERY"),
        )
        .select(["brand", "sub_channel", "date", "n_planned_promos", "avg_planned_discount"])
    )


def attach_event_importance(monthly: pl.DataFrame) -> pl.DataFrame:
    """Add per-month event-importance signal.

    `uk_holidays_count` only captures *number* of bank holidays — it can't
    tell the model that a month contains a World Cup final vs an ordinary
    August. This feature surfaces the strongest curated event landing in a
    given month as a numeric score (0 = none, 1 = LOW, 2 = MEDIUM, 3 = HIGH)
    plus three one-hot columns so LightGBM can pick up any non-linear
    interactions (e.g. "World Cup × hot weather"):

        event_importance_score : 0..3
        event_high  : 0/1
        event_med   : 0/1
        event_low   : 0/1

    Historical data shows HIGH-event months sell ~33% more on average than
    quiet months for Estrella × GROCERY, so this is a real signal the
    forecast was previously leaving on the table.
    """
    from app.services.calendar import build_events

    dmin, dmax = monthly["date"].min(), monthly["date"].max()
    events = build_events(dmin, dmax)

    # Strongest event per month → numeric score + one-hot. Multiple events in
    # the same month collapse to the highest importance (no stacking).
    rank = {"high": 3, "medium": 2, "low": 1}
    best_per_month: dict[date, int] = {}
    for e in events:
        # period is "YYYY-MM-DD" string for the month-start.
        from datetime import datetime as _dt
        m = _dt.fromisoformat(e.period).date()
        score = rank.get(e.importance, 0)
        if score > best_per_month.get(m, 0):
            best_per_month[m] = score

    out = monthly.with_columns(
        pl.col("date").replace_strict(
            best_per_month, default=0,
            return_dtype=pl.Int32,
        ).alias("event_importance_score")
    ).with_columns(
        (pl.col("event_importance_score") == 3).cast(pl.Int8).alias("event_high"),
        (pl.col("event_importance_score") == 2).cast(pl.Int8).alias("event_med"),
        (pl.col("event_importance_score") == 1).cast(pl.Int8).alias("event_low"),
    )
    n_high = int(out.filter(pl.col("event_high") == 1).height)
    n_med = int(out.filter(pl.col("event_med") == 1).height)
    print(f"  · attached event importance: {n_high} HIGH-event month-rows · {n_med} MEDIUM")
    return out


# ────────────────────────────────────────────────────────────────────────────
# Validation
# ────────────────────────────────────────────────────────────────────────────

def validate_monthly(monthly: pl.DataFrame) -> None:
    assert len(monthly) > 0, "monthly is empty"
    assert monthly["Hl"].min() > 0, "non-positive Hl in monthly"
    sub_set = set(monthly["sub_channel"].unique().to_list())
    bad = sub_set - ALLOWED_SUB_CHANNELS
    assert not bad, f"unexpected sub_channels: {bad}"
    assert monthly["date"].null_count() == 0, "null dates in monthly"
    print(f"  · monthly validation passed ({len(monthly):,} rows)")


def validate_promos(promos: pl.DataFrame) -> None:
    if len(promos) == 0:
        print("  ! promos is empty — parser found nothing")
        return
    bad = set(promos["promo_type"].unique().to_list()) - PROMO_TYPES
    assert not bad, f"unexpected promo_type: {bad}"
    assert promos["channel"].null_count() == 0, "null channel in promos"
    by_type = dict(promos.group_by("promo_type").agg(pl.len().alias("n")).iter_rows())
    by_channel = dict(promos.group_by("channel").agg(pl.len().alias("n")).iter_rows())
    print(f"  · promo types: {by_type}")
    print(f"  · per channel: {by_channel}")


def validate_targets(targets: pl.DataFrame, monthly: pl.DataFrame) -> None:
    assert len(targets) == len(monthly), \
        f"targets/monthly row mismatch: {len(targets)} vs {len(monthly)}"
    assert targets["target_hl"].null_count() == 0, "null target_hl after coalesce"
    assert set(targets["target_source"].unique().to_list()) <= {"prior_year", "trailing_median"}


# ────────────────────────────────────────────────────────────────────────────
# Output writers
# ────────────────────────────────────────────────────────────────────────────

def write_snapshots(monthly: pl.DataFrame, targets: pl.DataFrame, promos: pl.DataFrame) -> None:
    SNAPSHOTS.mkdir(parents=True, exist_ok=True)
    monthly.write_parquet(SNAPSHOTS / "wide_monthly.parquet")
    targets.write_parquet(SNAPSHOTS / "targets.parquet")
    promos.write_parquet(SNAPSHOTS / "promos.parquet")
    print(f"  · wrote wide_monthly.parquet ({len(monthly):,} rows)")
    print(f"  · wrote targets.parquet      ({len(targets):,} rows)")
    print(f"  · wrote promos.parquet       ({len(promos):,} rows)")


def write_meta(monthly: pl.DataFrame) -> None:
    brands = sorted(monthly["brand"].drop_nulls().unique().to_list())
    sub_channels = sorted(monthly["sub_channel"].drop_nulls().unique().to_list())
    sales_channels = sorted(monthly["sales_channel"].drop_nulls().unique().to_list())
    # Pull SKU labels from a fresh MaterialData load (they aren't in the monthly aggregate)
    mat = load_materials().select(["material_id", "label"])
    skus = (
        monthly.group_by("material_id")
        .agg(pl.col("brand").first(), pl.col("Hl").sum().alias("total_hl"))
        .sort("total_hl", descending=True)
        .join(mat, on="material_id", how="left")
        .with_columns(label=pl.coalesce(["label", "material_id"]))
        .select(
            pl.col("material_id").alias("id"),
            "label",
            "brand",
        )
        .to_dicts()
    )
    periods = sorted(monthly["date"].unique().to_list())
    period_strings = [d.strftime("%Y-%m") for d in periods]

    top_brand = monthly.group_by("brand").agg(pl.col("Hl").sum()).sort("Hl", descending=True)[0, "brand"]
    hero_row = (
        monthly.filter((pl.col("brand") == top_brand) & (pl.col("sub_channel") == "GROCERY"))
        .group_by("material_id")
        .agg(pl.col("Hl").sum().alias("total_hl"))
        .sort("total_hl", descending=True)
        .head(1)
    )
    hero = {
        "sku": hero_row[0, "material_id"] if len(hero_row) else None,
        "brand": top_brand, "sub_channel": "GROCERY", "period": period_strings[-1],
    }

    # Add human-readable channel labels via anonymize layer
    from app.services.anonymize import sub_channel_label, sales_channel_label
    sub_channels_labeled = [
        {"code": s, "label": sub_channel_label(s)} for s in sub_channels
    ]
    sales_channels_labeled = [
        {"code": s, "label": sales_channel_label(s)} for s in sales_channels
    ]

    meta = {
        "brands": brands, "skus": skus,
        "sub_channels": sub_channels,                       # raw codes (back-compat)
        "sub_channels_labeled": sub_channels_labeled,       # [{code, label}]
        "sales_channels": sales_channels,
        "sales_channels_labeled": sales_channels_labeled,
        "period_range": [period_strings[0], period_strings[-1]],
        "n_months": len(periods), "n_skus": len(skus),
        "hero": hero,
    }
    (SNAPSHOTS / "meta.json").write_text(json.dumps(meta, indent=2, default=str))
    print(f"  · wrote meta.json ({len(brands)} brands, {len(skus)} SKUs, {meta['period_range']})")


# ────────────────────────────────────────────────────────────────────────────
# CLI entry point
# ────────────────────────────────────────────────────────────────────────────

def main() -> int:
    print("=" * 70)
    print("MarketPulse UK — ETL")
    print("=" * 70)

    print("\n[1/8] Loading sales (DATABASE sheet)")
    sales_raw = load_sales_raw()

    print("\n[2/8] Loading UK customers")
    customers_uk = load_customers_uk()

    print("\n[3/8] Loading materials")
    materials = load_materials()

    print("\n[4/8] Filtering to actuals (null-Hl rows are accounting noise, not budget)")
    actuals = filter_actuals(sales_raw)

    print("\n[5/8] Netting returns")
    actuals_netted = net_returns(actuals)

    print("\n[6/8] Joining sales × customers × materials (UK only)")
    joined = join_uk(actuals_netted, customers_uk, materials)

    print("\n[7/9] Parsing promos (needed before monthly enrichment)")
    promos = parse_promos_all()
    validate_promos(promos)

    print("\n[8/9] Aggregating to monthly grain + holidays + external + planned promos + events")
    monthly = aggregate_monthly(joined)
    monthly = attach_uk_holidays(monthly)
    monthly = attach_external(monthly)
    monthly = attach_planned_promos(monthly, promos)
    monthly = attach_event_importance(monthly)
    validate_monthly(monthly)
    targets = derive_targets(monthly)
    validate_targets(targets, monthly)

    # Also generate future targets for the forecast horizon (9 months)
    future_targets = derive_future_targets(monthly, horizon_months=9)
    targets = pl.concat([targets, future_targets], how="vertical")
    print(f"  · future targets: +{len(future_targets):,} rows; total targets: {len(targets):,}")

    print("\n[9/9] Writing snapshots")
    write_snapshots(monthly, targets, promos)
    write_meta(monthly)

    print("\nDone.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
